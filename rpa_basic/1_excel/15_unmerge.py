from openpyxl import Workbook, load_workbook
wb = load_workbook("sample_merge.xlsx")
ws = wb.active

#B2:D2 병합되어 있던 셀을 해제
ws.unmerge_cells("B2:D2") #B2 부터 D2까지 
wb.save("sample_unmerge.xlsx")