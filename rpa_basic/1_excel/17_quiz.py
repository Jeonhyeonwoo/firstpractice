from openpyxl import Workbook

wb = Workbook()
ws = wb.active

#학점 출석 퀴즈1 퀴즈2 중간고사 기말고사 프로젝트 총점 성적  (퀴즈2 점수는 전부다 10점 )
# 90이상 A 80이상 B 70이상 C 나머지는 D 출석 점수가 5점미만이면 무조건 F 
ws.append(("학번","출석","퀴즈1","퀴즈2","중간고사","기말고사","프로젝트","총점","성적"))
scores = [
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)]

for i in scores:
    ws.append(i)

for idx , cell in enumerate(ws["D"]):
    if idx == 0:
        continue
    else:
        cell.value = 10

for idx, score in enumerate(scores, start=2):
    # sum_val = sum(score[1:]) - score[3] + 10 -> 이렇게 이용해서 A B C D 얻어도 됨 
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx ,idx)
    if score[1] < 5:
        ws.cell(row=idx,column=9).value = "F"
    else:
        ws.cell(row=idx, column=9).value = "=IF( H{}<70, \"D\" , IF( H{}<80,\"C\" , IF( H{}<90 ,\"B\", \"A\") ) )".format(idx,idx,idx)




wb.save("score.xlsx")