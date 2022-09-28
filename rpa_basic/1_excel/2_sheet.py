from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #새로운 sheet 기본 이름으로 생성
ws.title="Mysheet" #sheet 이름 변경
ws.sheet_properties.tabcolor = "ff66ff" #RGB 형태로 값을 넣어주면 택 색상 변경

ws1 = wb.create_sheet("Yoursheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("Newsheet",2) # 2번째 index 에 sheet 생성  

new_ws = wb["Newsheet"] # Dict 형태로 sheet 에 접근

print(wb.sheetnames) #모든 sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied sheet"

wb.save("sample.xlsx")